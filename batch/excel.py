from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from batch.models import ExcelRow


REQUIRED_HEADERS = ["voiceover_text", "emotion", "activity_name", "voiceover_title"]


class ExcelReaderError(ValueError):
    pass


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def read_excel_rows(path: str | Path) -> list[ExcelRow]:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    rows_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ExcelReaderError("Excel file is empty.") from exc

    headers = [_normalize_cell(cell) for cell in header_row]
    if headers[: len(REQUIRED_HEADERS)] != REQUIRED_HEADERS:
        raise ExcelReaderError(
            "Invalid headers. Expected first columns exactly: "
            + ", ".join(REQUIRED_HEADERS)
        )

    header_positions = {name: idx for idx, name in enumerate(headers)}
    parsed_rows: list[ExcelRow] = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        voiceover_text = _normalize_cell(row[header_positions["voiceover_text"]]) if len(row) > header_positions["voiceover_text"] else ""
        if not voiceover_text:
            continue

        emotion = _normalize_cell(row[header_positions["emotion"]]) if len(row) > header_positions["emotion"] else ""
        activity_name = (
            _normalize_cell(row[header_positions["activity_name"]])
            if len(row) > header_positions["activity_name"]
            else ""
        )
        voiceover_title = _normalize_cell(row[header_positions["voiceover_title"]]) if len(row) > header_positions["voiceover_title"] else ""

        parsed_rows.append(
            ExcelRow(
                row_index=idx,
                voiceover_text=voiceover_text,
                emotion=emotion,
                activity_name=activity_name,
                voiceover_title=voiceover_title,
            )
        )

    return parsed_rows
