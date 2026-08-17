from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


class BatchExcelError(ValueError):
    pass


@dataclass(frozen=True)
class HeyGenBatchRow:
    row_index: int
    script: str
    video_title: str


def _normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_heygen_batch_rows_json(payload: str) -> list[HeyGenBatchRow]:
    """Parse batch rows sent as JSON instead of a spreadsheet.

    The second way into the same batch: generated scripts arrive already
    reviewed and edited in the browser, so there is nothing to put in a file.
    Deliberately produces the identical ``HeyGenBatchRow`` list the Excel reader
    does — the runner cannot tell which way a batch was submitted, and neither
    path can grow behaviour the other lacks.

    Row indices are assigned here rather than trusted from the caller: they
    identify a row for the whole life of the batch, and duplicates from a client
    would collide in the batch store's per-row state.
    """
    try:
        items = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise BatchExcelError(f"rows is not valid JSON: {exc}") from exc

    if not isinstance(items, list):
        raise BatchExcelError("rows must be a JSON array of {script, video_title} objects")

    rows: list[HeyGenBatchRow] = []
    for idx, item in enumerate(items, start=2):
        if not isinstance(item, dict):
            raise BatchExcelError("each row must be an object with script and video_title")
        script = _normalize(item.get("script"))
        if not script:
            continue
        video_title = _normalize(item.get("video_title")) or f"row_{idx}"
        rows.append(HeyGenBatchRow(row_index=idx, script=script, video_title=video_title))

    return rows


def read_heygen_batch_rows(path: str | Path) -> list[HeyGenBatchRow]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration as exc:
        raise BatchExcelError("Excel file is empty.") from exc

    headers = [_normalize(c).lower() for c in header_row]

    if "script" not in headers:
        raise BatchExcelError(
            "Missing required column 'script'. Expected header row: script, video_title"
        )

    script_idx = headers.index("script")
    title_idx = headers.index("video_title") if "video_title" in headers else None

    rows: list[HeyGenBatchRow] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        script = _normalize(row[script_idx]) if len(row) > script_idx else ""
        if not script:
            continue
        video_title = ""
        if title_idx is not None and len(row) > title_idx:
            video_title = _normalize(row[title_idx])
        if not video_title:
            video_title = f"row_{idx}"
        rows.append(HeyGenBatchRow(row_index=idx, script=script, video_title=video_title))

    return rows
